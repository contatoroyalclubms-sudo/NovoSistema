from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from typing import List, Optional
from datetime import date, datetime, timedelta
from decimal import Decimal
import uuid
import os
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

from ..database import get_db
from ..models import (
    MovimentacaoFinanceira, CaixaEvento, Evento, Usuario, 
    TipoMovimentacaoFinanceira, StatusMovimentacaoFinanceira,
    Transacao, VendaPDV, LogAuditoria
)
from ..schemas import (
    MovimentacaoFinanceiraCreate, MovimentacaoFinanceiraUpdate, 
    MovimentacaoFinanceira as MovimentacaoFinanceiraSchema,
    CaixaEventoCreate, CaixaEvento as CaixaEventoSchema,
    DashboardFinanceiro
)
from ..auth import obter_usuario_atual, verificar_permissao_admin, verificar_permissao_promoter

router = APIRouter(prefix="/financeiro", tags=["Financeiro"])

@router.post("/movimentacoes", response_model=MovimentacaoFinanceiraSchema)
async def criar_movimentacao(
    movimentacao: MovimentacaoFinanceiraCreate,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_promoter)
):
    """Criar nova movimentação financeira"""
    
    evento = db.query(Evento).filter(Evento.id == movimentacao.evento_id).first()
    if not evento:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    db_movimentacao = MovimentacaoFinanceira(
        **movimentacao.dict(),
        usuario_responsavel_id=usuario_atual.id
    )
    
    db.add(db_movimentacao)
    db.commit()
    db.refresh(db_movimentacao)
    
    log = LogAuditoria(
        cpf_usuario=usuario_atual.cpf,
        acao="criar_movimentacao_financeira",
        tabela_afetada="movimentacoes_financeiras",
        registro_id=db_movimentacao.id,
        evento_id=movimentacao.evento_id,
        dados_novos=f"Tipo: {movimentacao.tipo}, Valor: {movimentacao.valor}",
        status="sucesso"
    )
    db.add(log)
    db.commit()
    
    return db_movimentacao

@router.get("/movimentacoes/{evento_id}", response_model=List[MovimentacaoFinanceiraSchema])
async def listar_movimentacoes(
    evento_id: int,
    tipo: Optional[str] = "",
    categoria: Optional[str] = "",
    data_inicio: Optional[str] = "",
    data_fim: Optional[str] = "",
    status: Optional[str] = "",
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Listar movimentações financeiras do evento"""
    
    evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not evento:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    query = db.query(MovimentacaoFinanceira).filter(
        MovimentacaoFinanceira.evento_id == evento_id
    )
    
    if tipo and tipo.strip():
        query = query.filter(MovimentacaoFinanceira.tipo == tipo)
    if categoria and categoria.strip():
        query = query.filter(MovimentacaoFinanceira.categoria.ilike(f"%{categoria}%"))
    if data_inicio and data_inicio.strip():
        try:
            data_inicio_parsed = datetime.strptime(data_inicio, "%Y-%m-%d").date()
            query = query.filter(MovimentacaoFinanceira.criado_em >= data_inicio_parsed)
        except ValueError:
            pass
    if data_fim and data_fim.strip():
        try:
            data_fim_parsed = datetime.strptime(data_fim, "%Y-%m-%d").date()
            query = query.filter(MovimentacaoFinanceira.criado_em <= data_fim_parsed + timedelta(days=1))
        except ValueError:
            pass
    if status and status.strip():
        query = query.filter(MovimentacaoFinanceira.status == status)
    
    movimentacoes = query.order_by(MovimentacaoFinanceira.criado_em.desc()).all()
    
    return movimentacoes

@router.put("/movimentacoes/{movimentacao_id}", response_model=MovimentacaoFinanceiraSchema)
async def atualizar_movimentacao(
    movimentacao_id: int,
    movimentacao_update: MovimentacaoFinanceiraUpdate,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_promoter)
):
    """Atualizar movimentação financeira"""
    
    movimentacao = db.query(MovimentacaoFinanceira).filter(
        MovimentacaoFinanceira.id == movimentacao_id
    ).first()
    
    if not movimentacao:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    
    evento = db.query(Evento).filter(Evento.id == movimentacao.evento_id).first()
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    dados_anteriores = {
        "categoria": movimentacao.categoria,
        "descricao": movimentacao.descricao,
        "valor": str(movimentacao.valor),
        "status": movimentacao.status.value
    }
    
    for field, value in movimentacao_update.dict(exclude_unset=True).items():
        setattr(movimentacao, field, value)
    
    db.commit()
    db.refresh(movimentacao)
    
    log = LogAuditoria(
        cpf_usuario=usuario_atual.cpf,
        acao="atualizar_movimentacao_financeira",
        tabela_afetada="movimentacoes_financeiras",
        registro_id=movimentacao.id,
        evento_id=movimentacao.evento_id,
        dados_anteriores=str(dados_anteriores),
        dados_novos=str(movimentacao_update.dict(exclude_unset=True)),
        status="sucesso"
    )
    db.add(log)
    db.commit()
    
    return movimentacao

@router.post("/movimentacoes/{movimentacao_id}/comprovante")
async def upload_comprovante(
    movimentacao_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_promoter)
):
    """Upload de comprovante para movimentação"""
    
    movimentacao = db.query(MovimentacaoFinanceira).filter(
        MovimentacaoFinanceira.id == movimentacao_id
    ).first()
    
    if not movimentacao:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    
    evento = db.query(Evento).filter(Evento.id == movimentacao.evento_id).first()
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    allowed_types = ["image/jpeg", "image/png", "application/pdf"]
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400, 
            detail="Tipo de arquivo não permitido. Use JPEG, PNG ou PDF."
        )
    
    upload_dir = "uploads/comprovantes"
    os.makedirs(upload_dir, exist_ok=True)
    
    file_extension = file.filename.split(".")[-1]
    filename = f"comprovante_{movimentacao_id}_{uuid.uuid4().hex}.{file_extension}"
    file_path = os.path.join(upload_dir, filename)
    
    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)
    
    movimentacao.comprovante_url = file_path
    db.commit()
    
    return {"message": "Comprovante enviado com sucesso", "url": file_path}

@router.get("/dashboard/{evento_id}", response_model=DashboardFinanceiro)
async def obter_dashboard_financeiro(
    evento_id: int,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Dashboard financeiro do evento"""
    
    evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not evento:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    total_entradas = db.query(func.sum(MovimentacaoFinanceira.valor)).filter(
        MovimentacaoFinanceira.evento_id == evento_id,
        MovimentacaoFinanceira.tipo == "entrada",
        MovimentacaoFinanceira.status == "aprovada"
    ).scalar() or Decimal('0.00')
    
    total_saidas = db.query(func.sum(MovimentacaoFinanceira.valor)).filter(
        MovimentacaoFinanceira.evento_id == evento_id,
        MovimentacaoFinanceira.tipo == "saida",
        MovimentacaoFinanceira.status == "aprovada"
    ).scalar() or Decimal('0.00')
    
    total_vendas_listas = db.query(func.sum(Transacao.valor)).filter(
        Transacao.evento_id == evento_id,
        Transacao.status == "aprovada"
    ).scalar() or Decimal('0.00')
    
    total_vendas_pdv = db.query(func.sum(VendaPDV.valor_final)).filter(
        VendaPDV.evento_id == evento_id,
        VendaPDV.status == "aprovada"
    ).scalar() or Decimal('0.00')
    
    total_vendas = total_vendas_listas + total_vendas_pdv
    saldo_atual = total_entradas + total_vendas - total_saidas
    lucro_prejuizo = saldo_atual
    
    caixa = db.query(CaixaEvento).filter(
        CaixaEvento.evento_id == evento_id,
        CaixaEvento.status == "aberto"
    ).first()
    
    status_caixa = "aberto" if caixa else "fechado"
    
    movimentacoes_recentes = db.query(MovimentacaoFinanceira).filter(
        MovimentacaoFinanceira.evento_id == evento_id
    ).order_by(MovimentacaoFinanceira.criado_em.desc()).limit(5).all()
    
    categorias_despesas = db.query(
        MovimentacaoFinanceira.categoria,
        func.sum(MovimentacaoFinanceira.valor).label('total')
    ).filter(
        MovimentacaoFinanceira.evento_id == evento_id,
        MovimentacaoFinanceira.tipo == "saida",
        MovimentacaoFinanceira.status == "aprovada"
    ).group_by(MovimentacaoFinanceira.categoria).all()
    
    repasses_promoters = db.query(
        Usuario.nome,
        func.sum(MovimentacaoFinanceira.valor).label('total')
    ).join(
        MovimentacaoFinanceira, MovimentacaoFinanceira.promoter_id == Usuario.id
    ).filter(
        MovimentacaoFinanceira.evento_id == evento_id,
        MovimentacaoFinanceira.tipo == "repasse_promoter",
        MovimentacaoFinanceira.status == "aprovada"
    ).group_by(Usuario.id, Usuario.nome).all()
    
    return DashboardFinanceiro(
        evento_id=evento_id,
        saldo_atual=saldo_atual,
        total_entradas=total_entradas,
        total_saidas=total_saidas,
        total_vendas=total_vendas,
        lucro_prejuizo=lucro_prejuizo,
        movimentacoes_recentes=[
            {
                "id": mov.id,
                "tipo": mov.tipo.value,
                "categoria": mov.categoria,
                "valor": float(mov.valor),
                "criado_em": mov.criado_em.isoformat()
            }
            for mov in movimentacoes_recentes
        ],
        categorias_despesas=[
            {"categoria": cat.categoria, "total": float(cat.total)}
            for cat in categorias_despesas
        ],
        repasses_promoters=[
            {"promoter": rep.nome, "total": float(rep.total)}
            for rep in repasses_promoters
        ],
        status_caixa=status_caixa
    )

@router.get("/relatorio/{evento_id}/export/{formato}")
async def exportar_relatorio_financeiro(
    evento_id: int,
    formato: str,
    data_inicio: Optional[str] = "",
    data_fim: Optional[str] = "",
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Exportar relatório financeiro em PDF, Excel ou CSV"""
    
    if formato not in ["pdf", "excel", "csv"]:
        raise HTTPException(status_code=400, detail="Formato não suportado")
    
    evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not evento:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    query = db.query(MovimentacaoFinanceira).filter(
        MovimentacaoFinanceira.evento_id == evento_id
    )
    
    if data_inicio and data_inicio.strip():
        try:
            data_inicio_parsed = datetime.strptime(data_inicio, "%Y-%m-%d").date()
            query = query.filter(MovimentacaoFinanceira.criado_em >= data_inicio_parsed)
        except ValueError:
            pass
    if data_fim and data_fim.strip():
        try:
            data_fim_parsed = datetime.strptime(data_fim, "%Y-%m-%d").date()
            query = query.filter(MovimentacaoFinanceira.criado_em <= data_fim_parsed + timedelta(days=1))
        except ValueError:
            pass
    
    movimentacoes = query.order_by(MovimentacaoFinanceira.criado_em.desc()).all()
    
    if formato == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório Financeiro"
        
        headers = ['Data', 'Tipo', 'Categoria', 'Descrição', 'Valor', 'Status', 'Responsável']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for row, mov in enumerate(movimentacoes, 2):
            ws.cell(row=row, column=1, value=mov.criado_em.strftime("%d/%m/%Y"))
            ws.cell(row=row, column=2, value=mov.tipo.value)
            ws.cell(row=row, column=3, value=mov.categoria)
            ws.cell(row=row, column=4, value=mov.descricao)
            ws.cell(row=row, column=5, value=float(mov.valor))
            ws.cell(row=row, column=6, value=mov.status.value)
            ws.cell(row=row, column=7, value=mov.usuario_responsavel.nome)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=financeiro_evento_{evento_id}.xlsx"}
        )
    
    elif formato == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['Data', 'Tipo', 'Categoria', 'Descrição', 'Valor', 'Status', 'Responsável'])
        
        for mov in movimentacoes:
            writer.writerow([
                mov.criado_em.strftime("%d/%m/%Y"),
                mov.tipo.value,
                mov.categoria,
                mov.descricao,
                str(mov.valor),
                mov.status.value,
                mov.usuario_responsavel.nome
            ])
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=financeiro_evento_{evento_id}.csv"}
        )
    
    elif formato == "pdf":
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, f"Relatório Financeiro - {evento.nome}")
        
        p.setFont("Helvetica", 12)
        y_position = height - 100
        
        total_entradas = sum(float(mov.valor) for mov in movimentacoes if mov.tipo.value == "entrada")
        total_saidas = sum(float(mov.valor) for mov in movimentacoes if mov.tipo.value == "saida")
        saldo = total_entradas - total_saidas
        
        p.drawString(50, y_position, f"Total Entradas: R$ {total_entradas:.2f}")
        y_position -= 20
        p.drawString(50, y_position, f"Total Saídas: R$ {total_saidas:.2f}")
        y_position -= 20
        p.drawString(50, y_position, f"Saldo: R$ {saldo:.2f}")
        y_position -= 40
        
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y_position, "Movimentações:")
        y_position -= 20
        
        p.setFont("Helvetica", 10)
        for mov in movimentacoes[:20]:
            if y_position < 50:
                p.showPage()
                y_position = height - 50
            
            linha = f"{mov.criado_em.strftime('%d/%m/%Y')} - {mov.tipo.value.upper()} - {mov.categoria} - R$ {float(mov.valor):.2f}"
            p.drawString(50, y_position, linha)
            y_position -= 15
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=financeiro_evento_{evento_id}.pdf"}
        )

@router.post("/caixa/abrir", response_model=CaixaEventoSchema)
async def abrir_caixa_evento(
    caixa: CaixaEventoCreate,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_promoter)
):
    """Abrir caixa do evento"""
    
    evento = db.query(Evento).filter(Evento.id == caixa.evento_id).first()
    if not evento:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    caixa_existente = db.query(CaixaEvento).filter(
        CaixaEvento.evento_id == caixa.evento_id,
        CaixaEvento.status == "aberto"
    ).first()
    
    if caixa_existente:
        raise HTTPException(status_code=400, detail="Já existe um caixa aberto para este evento")
    
    db_caixa = CaixaEvento(
        **caixa.dict(),
        usuario_abertura_id=usuario_atual.id
    )
    
    db.add(db_caixa)
    db.commit()
    db.refresh(db_caixa)
    
    log = LogAuditoria(
        cpf_usuario=usuario_atual.cpf,
        acao="abrir_caixa_evento",
        tabela_afetada="caixas_eventos",
        registro_id=db_caixa.id,
        evento_id=caixa.evento_id,
        dados_novos=f"Saldo inicial: {caixa.saldo_inicial}",
        status="sucesso"
    )
    db.add(log)
    db.commit()
    
    return db_caixa

@router.post("/caixa/{caixa_id}/fechar")
async def fechar_caixa_evento(
    caixa_id: int,
    observacoes_fechamento: Optional[str] = None,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_promoter)
):
    """Fechar caixa do evento"""
    
    caixa = db.query(CaixaEvento).filter(CaixaEvento.id == caixa_id).first()
    if not caixa:
        raise HTTPException(status_code=404, detail="Caixa não encontrado")
    
    evento = db.query(Evento).filter(Evento.id == caixa.evento_id).first()
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    if caixa.status == "fechado":
        raise HTTPException(status_code=400, detail="Caixa já está fechado")
    
    total_entradas = db.query(func.sum(MovimentacaoFinanceira.valor)).filter(
        MovimentacaoFinanceira.evento_id == caixa.evento_id,
        MovimentacaoFinanceira.tipo == "entrada",
        MovimentacaoFinanceira.status == "aprovada"
    ).scalar() or Decimal('0.00')
    
    total_saidas = db.query(func.sum(MovimentacaoFinanceira.valor)).filter(
        MovimentacaoFinanceira.evento_id == caixa.evento_id,
        MovimentacaoFinanceira.tipo == "saida",
        MovimentacaoFinanceira.status == "aprovada"
    ).scalar() or Decimal('0.00')
    
    total_vendas_listas = db.query(func.sum(Transacao.valor)).filter(
        Transacao.evento_id == caixa.evento_id,
        Transacao.status == "aprovada"
    ).scalar() or Decimal('0.00')
    
    total_vendas_pdv = db.query(func.sum(VendaPDV.valor_final)).filter(
        VendaPDV.evento_id == caixa.evento_id,
        VendaPDV.status == "aprovada"
    ).scalar() or Decimal('0.00')
    
    caixa.total_entradas = total_entradas
    caixa.total_saidas = total_saidas
    caixa.total_vendas_listas = total_vendas_listas
    caixa.total_vendas_pdv = total_vendas_pdv
    caixa.saldo_final = caixa.saldo_inicial + total_entradas + total_vendas_listas + total_vendas_pdv - total_saidas
    caixa.status = "fechado"
    caixa.data_fechamento = datetime.now()
    caixa.usuario_fechamento_id = usuario_atual.id
    caixa.observacoes_fechamento = observacoes_fechamento
    
    db.commit()
    
    log = LogAuditoria(
        cpf_usuario=usuario_atual.cpf,
        acao="fechar_caixa_evento",
        tabela_afetada="caixas_eventos",
        registro_id=caixa.id,
        evento_id=caixa.evento_id,
        dados_novos=f"Saldo final: {caixa.saldo_final}",
        status="sucesso"
    )
    db.add(log)
    db.commit()
    
    return {"message": "Caixa fechado com sucesso", "saldo_final": float(caixa.saldo_final)}

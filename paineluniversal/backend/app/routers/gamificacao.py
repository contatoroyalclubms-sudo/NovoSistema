from fastapi import APIRouter, Depends, HTTPException, status, Response, BackgroundTasks
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, and_, or_
from typing import List, Optional
from datetime import date, datetime, timedelta
from decimal import Decimal
import uuid
import os
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

from ..database import get_db
from ..models import (
    Usuario, Evento, Lista, Transacao, Checkin, PromoterEvento,
    Conquista, PromoterConquista, MetricaPromoter, TipoConquista, NivelBadge,
    LogAuditoria
)
from ..schemas import (
    ConquistaCreate, Conquista as ConquistaSchema,
    MetricaPromoterResponse, RankingGamificado, DashboardGamificacao,
    FiltrosRanking, PromoterConquistaResponse
)
from ..auth import obter_usuario_atual, verificar_permissao_admin, verificar_permissao_promoter
from ..services.whatsapp_service import whatsapp_service

router = APIRouter(prefix="/gamificacao", tags=["GamificaÃ§Ã£o"])

@router.get("/ranking", response_model=List[RankingGamificado])
async def obter_ranking_gamificado(
    evento_id: Optional[int] = None,
    periodo_inicio: Optional[date] = None,
    periodo_fim: Optional[date] = None,
    badge_nivel: Optional[str] = None,
    tipo_ranking: Optional[str] = "geral",
    limit: int = 20,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Obter ranking gamificado de promoters"""
    
    if not periodo_inicio:
        periodo_inicio = date.today() - timedelta(days=30)
    if not periodo_fim:
        periodo_fim = date.today()
    
    query = db.query(
        Usuario.id.label('promoter_id'),
        Usuario.nome.label('nome_promoter'),
        func.count(Transacao.id).label('total_vendas'),
        func.sum(Transacao.valor).label('receita_gerada'),
        func.count(Checkin.id).label('total_presentes'),
        func.count(Lista.id).label('total_listas')
    ).join(
        Lista, Lista.promoter_id == Usuario.id
    ).join(
        Transacao, Transacao.lista_id == Lista.id
    ).outerjoin(
        Checkin, and_(
            Transacao.cpf_comprador == Checkin.cpf,
            Transacao.evento_id == Checkin.evento_id
        )
    ).filter(
        Usuario.tipo == "promoter",
        Transacao.status == "aprovada",
        func.date(Transacao.criado_em).between(periodo_inicio, periodo_fim)
    )
    
    if usuario_atual.tipo.value != "admin":
        query = query.join(Evento).filter(Evento.empresa_id == usuario_atual.empresa_id)
    
    if evento_id:
        query = query.filter(Transacao.evento_id == evento_id)
    
    resultados = query.group_by(
        Usuario.id, Usuario.nome
    ).order_by(
        desc(func.sum(Transacao.valor))
    ).limit(limit).all()
    
    ranking = []
    for i, resultado in enumerate(resultados, 1):
        taxa_presenca = (resultado.total_presentes / resultado.total_vendas * 100) if resultado.total_vendas > 0 else 0
        
        conquistas_count = db.query(func.count(PromoterConquista.id)).filter(
            PromoterConquista.promoter_id == resultado.promoter_id
        ).scalar() or 0
        
        conquistas_mes = db.query(func.count(PromoterConquista.id)).filter(
            PromoterConquista.promoter_id == resultado.promoter_id,
            func.date(PromoterConquista.data_conquista) >= periodo_inicio
        ).scalar() or 0
        
        badge_principal = calcular_badge_principal(i, resultado.total_vendas, taxa_presenca)
        
        pontuacao = calcular_pontuacao_gamificada(
            resultado.total_vendas,
            float(resultado.receita_gerada or 0),
            taxa_presenca,
            conquistas_count
        )
        
        ranking.append(RankingGamificado(
            promoter_id=resultado.promoter_id,
            nome_promoter=resultado.nome_promoter,
            badge_principal=badge_principal,
            nivel_experiencia=min(100, conquistas_count * 5),
            total_vendas=resultado.total_vendas,
            receita_gerada=resultado.receita_gerada or Decimal('0.00'),
            taxa_presenca=round(taxa_presenca, 2),
            taxa_conversao=round(taxa_presenca, 2),
            crescimento_mensal=0.0,
            posicao_atual=i,
            posicao_anterior=None,
            conquistas_total=conquistas_count,
            conquistas_mes=conquistas_mes,
            eventos_ativos=resultado.total_listas,
            streak_vendas=0,
            pontuacao_total=pontuacao
        ))
    
    return ranking

@router.get("/dashboard", response_model=DashboardGamificacao)
async def obter_dashboard_gamificacao(
    evento_id: Optional[int] = None,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Dashboard completo de gamificaÃ§Ã£o"""
    
    ranking_geral = await obter_ranking_gamificado(
        evento_id=evento_id, limit=10, db=db, usuario_atual=usuario_atual
    )
    
    conquistas_recentes = db.query(
        PromoterConquista.id,
        Conquista.nome.label('conquista_nome'),
        Conquista.descricao.label('conquista_descricao'),
        Conquista.badge_nivel,
        Conquista.icone,
        PromoterConquista.valor_alcancado,
        PromoterConquista.data_conquista,
        Evento.nome.label('evento_nome')
    ).join(
        Conquista, PromoterConquista.conquista_id == Conquista.id
    ).join(
        Usuario, PromoterConquista.promoter_id == Usuario.id
    ).outerjoin(
        Evento, PromoterConquista.evento_id == Evento.id
    ).filter(
        PromoterConquista.data_conquista >= datetime.now() - timedelta(days=7)
    )
    
    if usuario_atual.tipo.value != "admin":
        conquistas_recentes = conquistas_recentes.filter(
            Usuario.empresa_id == usuario_atual.empresa_id
        )
    
    conquistas_recentes = conquistas_recentes.order_by(
        desc(PromoterConquista.data_conquista)
    ).limit(20).all()
    
    conquistas_list = [
        PromoterConquistaResponse(
            id=c.id,
            conquista_nome=c.conquista_nome,
            conquista_descricao=c.conquista_descricao,
            badge_nivel=c.badge_nivel.value,
            icone=c.icone,
            valor_alcancado=c.valor_alcancado,
            data_conquista=c.data_conquista,
            evento_nome=c.evento_nome
        ) for c in conquistas_recentes
    ]
    
    return DashboardGamificacao(
        ranking_geral=ranking_geral,
        conquistas_recentes=conquistas_list,
        metricas_periodo={
            "total_promoters": len(ranking_geral),
            "conquistas_semana": len(conquistas_list),
            "badge_ouro": len([r for r in ranking_geral if r.badge_principal == "ouro"]),
            "crescimento_medio": 15.5
        },
        badges_disponiveis=[
            {"nome": "Bronze", "descricao": "10+ vendas", "icone": "ðŸ¥‰"},
            {"nome": "Prata", "descricao": "50+ vendas", "icone": "ðŸ¥ˆ"},
            {"nome": "Ouro", "descricao": "100+ vendas", "icone": "ðŸ¥‡"},
            {"nome": "Platina", "descricao": "200+ vendas + 80% presenÃ§a", "icone": "ðŸ’Ž"},
            {"nome": "Diamante", "descricao": "500+ vendas + 90% presenÃ§a", "icone": "ðŸ’ "},
            {"nome": "Lenda", "descricao": "1000+ vendas + mÃºltiplas conquistas", "icone": "ðŸ‘‘"}
        ],
        alertas_gamificacao=[],
        estatisticas_gerais={
            "total_badges_distribuidos": sum(r.conquistas_total for r in ranking_geral),
            "promoter_mais_ativo": ranking_geral[0].nome_promoter if ranking_geral else None,
            "meta_mensal": "80% dos promoters com badge prata+"
        }
    )

@router.post("/conquistas", response_model=ConquistaSchema)
async def criar_conquista(
    conquista: ConquistaCreate,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_admin)
):
    """Criar nova conquista (apenas admin)"""
    
    db_conquista = Conquista(**conquista.dict())
    db.add(db_conquista)
    db.commit()
    db.refresh(db_conquista)
    
    log = LogAuditoria(
        cpf_usuario=usuario_atual.cpf,
        acao="criar_conquista",
        tabela_afetada="conquistas",
        registro_id=db_conquista.id,
        dados_novos=f"Nome: {conquista.nome}, Tipo: {conquista.tipo}",
        status="sucesso"
    )
    db.add(log)
    db.commit()
    
    return db_conquista

@router.post("/verificar-conquistas/{promoter_id}")
async def verificar_conquistas_promoter(
    promoter_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_admin)
):
    """Verificar e atribuir conquistas para um promoter"""
    
    promoter = db.query(Usuario).filter(
        Usuario.id == promoter_id,
        Usuario.tipo == "promoter"
    ).first()
    
    if not promoter:
        raise HTTPException(status_code=404, detail="Promoter nÃ£o encontrado")
    
    conquistas_disponiveis = db.query(Conquista).filter(Conquista.ativa == True).all()
    
    novas_conquistas = []
    
    for conquista in conquistas_disponiveis:
        ja_possui = db.query(PromoterConquista).filter(
            PromoterConquista.promoter_id == promoter_id,
            PromoterConquista.conquista_id == conquista.id
        ).first()
        
        if ja_possui:
            continue
        
        valor_alcancado = 0
        
        if conquista.tipo == TipoConquista.VENDAS:
            valor_alcancado = db.query(func.count(Transacao.id)).join(Lista).filter(
                Lista.promoter_id == promoter_id,
                Transacao.status == "aprovada"
            ).scalar() or 0
        
        elif conquista.tipo == TipoConquista.PRESENCA:
            total_vendas = db.query(func.count(Transacao.id)).join(Lista).filter(
                Lista.promoter_id == promoter_id,
                Transacao.status == "aprovada"
            ).scalar() or 0
            
            total_presentes = db.query(func.count(Checkin.id)).join(
                Transacao, Transacao.cpf_comprador == Checkin.cpf
            ).join(Lista).filter(
                Lista.promoter_id == promoter_id
            ).scalar() or 0
            
            valor_alcancado = int((total_presentes / total_vendas * 100)) if total_vendas > 0 else 0
        
        if valor_alcancado >= conquista.criterio_valor:
            nova_conquista = PromoterConquista(
                promoter_id=promoter_id,
                conquista_id=conquista.id,
                valor_alcancado=valor_alcancado
            )
            db.add(nova_conquista)
            novas_conquistas.append(conquista)
    
    if novas_conquistas:
        db.commit()
        
        if promoter.telefone:
            background_tasks.add_task(
                enviar_notificacao_conquista,
                promoter.telefone,
                promoter.nome,
                novas_conquistas
            )
    
    return {
        "message": f"{len(novas_conquistas)} novas conquistas atribuÃ­das",
        "conquistas": [c.nome for c in novas_conquistas]
    }

@router.get("/export/ranking/{formato}")
async def exportar_ranking(
    formato: str,
    evento_id: Optional[int] = None,
    periodo_inicio: Optional[date] = None,
    periodo_fim: Optional[date] = None,
    badge_nivel: Optional[str] = None,
    tipo_ranking: Optional[str] = "geral",
    limit: int = 20,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Exportar ranking em Excel, PDF ou CSV"""
    
    if formato not in ["excel", "pdf", "csv"]:
        raise HTTPException(status_code=400, detail="Formato nÃ£o suportado")
    
    ranking = await obter_ranking_gamificado(
        evento_id=evento_id,
        periodo_inicio=periodo_inicio,
        periodo_fim=periodo_fim,
        badge_nivel=badge_nivel,
        tipo_ranking=tipo_ranking,
        limit=limit,
        db=db,
        usuario_atual=usuario_atual
    )
    
    if formato == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Ranking Promoters"
        
        headers = [
            'PosiÃ§Ã£o', 'Nome', 'Badge', 'Vendas', 'Receita', 'Taxa PresenÃ§a (%)',
            'Conquistas', 'PontuaÃ§Ã£o', 'NÃ­vel'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row, promoter in enumerate(ranking, 2):
            ws.cell(row=row, column=1, value=promoter.posicao_atual)
            ws.cell(row=row, column=2, value=promoter.nome_promoter)
            ws.cell(row=row, column=3, value=promoter.badge_principal.upper())
            ws.cell(row=row, column=4, value=promoter.total_vendas)
            ws.cell(row=row, column=5, value=float(promoter.receita_gerada))
            ws.cell(row=row, column=6, value=promoter.taxa_presenca)
            ws.cell(row=row, column=7, value=promoter.conquistas_total)
            ws.cell(row=row, column=8, value=promoter.pontuacao_total)
            ws.cell(row=row, column=9, value=promoter.nivel_experiencia)
        
        chart = BarChart()
        chart.title = "Top 10 Promoters - Vendas"
        chart.x_axis.title = "Promoters"
        chart.y_axis.title = "Vendas"
        
        data = Reference(ws, min_col=4, min_row=1, max_row=min(11, len(ranking) + 1))
        categories = Reference(ws, min_col=2, min_row=2, max_row=min(11, len(ranking) + 1))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        ws.add_chart(chart, "K2")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=ranking_promoters.xlsx"}
        )
    
    elif formato == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow([
            'PosiÃ§Ã£o', 'Nome', 'Badge', 'Vendas', 'Receita', 'Taxa PresenÃ§a (%)',
            'Conquistas', 'PontuaÃ§Ã£o', 'NÃ­vel'
        ])
        
        for promoter in ranking:
            writer.writerow([
                promoter.posicao_atual,
                promoter.nome_promoter,
                promoter.badge_principal.upper(),
                promoter.total_vendas,
                str(promoter.receita_gerada),
                promoter.taxa_presenca,
                promoter.conquistas_total,
                promoter.pontuacao_total,
                promoter.nivel_experiencia
            ])
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=ranking_promoters.csv"}
        )
    
    return {"message": "Formato PDF em desenvolvimento"}

def calcular_badge_principal(posicao: int, vendas: int, taxa_presenca: float) -> str:
    """Calcular badge principal baseado em mÃ©tricas"""
    if vendas >= 1000 and taxa_presenca >= 90:
        return "lenda"
    elif vendas >= 500 and taxa_presenca >= 90:
        return "diamante"
    elif vendas >= 200 and taxa_presenca >= 80:
        return "platina"
    elif vendas >= 100:
        return "ouro"
    elif vendas >= 50:
        return "prata"
    else:
        return "bronze"

def calcular_pontuacao_gamificada(vendas: int, receita: float, taxa_presenca: float, conquistas: int) -> int:
    """Calcular pontuaÃ§Ã£o gamificada total"""
    pontos_vendas = vendas * 10
    pontos_receita = int(receita / 10)
    pontos_presenca = int(taxa_presenca * 5)
    pontos_conquistas = conquistas * 100
    
    return pontos_vendas + pontos_receita + pontos_presenca + pontos_conquistas

async def enviar_notificacao_conquista(telefone: str, nome: str, conquistas: List[Conquista]):
    """Enviar notificaÃ§Ã£o de conquista via WhatsApp"""
    conquistas_texto = "\n".join([f"{c.icone} {c.nome}" for c in conquistas])
    
    message = f"""
ðŸŽ‰ *PARABÃ‰NS {nome.upper()}!*

VocÃª conquistou novos badges:
{conquistas_texto}

Continue assim e alcance novos nÃ­veis! ðŸš€
    """.strip()
    
    await whatsapp_service._send_whatsapp_message(telefone, message)

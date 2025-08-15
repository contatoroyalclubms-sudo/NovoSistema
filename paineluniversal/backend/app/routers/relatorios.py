from fastapi import APIRouter, Depends, HTTPException, status, Response
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime, date
from ..database import get_db
from ..models import Evento, Transacao, Checkin, Usuario, Lista
from ..schemas import RelatorioVendas
from ..auth import obter_usuario_atual, verificar_permissao_admin
import csv
import io
import json
from decimal import Decimal

router = APIRouter()

@router.get("/vendas/{evento_id}", response_model=RelatorioVendas)
async def gerar_relatorio_vendas(
    evento_id: int,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Gerar relatório de vendas de um evento"""
    
    evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not evento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento não encontrado"
        )
    
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acesso negado"
        )
    
    transacoes = db.query(Transacao).filter(
        Transacao.evento_id == evento_id,
        Transacao.status == "aprovada"
    ).all()
    
    total_vendas = len(transacoes)
    receita_total = sum(t.valor for t in transacoes)
    
    vendas_por_lista = {}
    for transacao in transacoes:
        lista = db.query(Lista).filter(Lista.id == transacao.lista_id).first()
        if lista:
            tipo_lista = lista.tipo.value
            if tipo_lista not in vendas_por_lista:
                vendas_por_lista[tipo_lista] = {"vendas": 0, "receita": Decimal('0.00')}
            vendas_por_lista[tipo_lista]["vendas"] += 1
            vendas_por_lista[tipo_lista]["receita"] += transacao.valor
    
    vendas_por_promoter = {}
    for transacao in transacoes:
        lista = db.query(Lista).filter(Lista.id == transacao.lista_id).first()
        if lista and lista.promoter_id:
            promoter = db.query(Usuario).filter(Usuario.id == lista.promoter_id).first()
            if promoter:
                nome_promoter = promoter.nome
                if nome_promoter not in vendas_por_promoter:
                    vendas_por_promoter[nome_promoter] = {"vendas": 0, "receita": Decimal('0.00')}
                vendas_por_promoter[nome_promoter]["vendas"] += 1
                vendas_por_promoter[nome_promoter]["receita"] += transacao.valor
    
    return RelatorioVendas(
        evento_id=evento_id,
        nome_evento=evento.nome,
        total_vendas=total_vendas,
        receita_total=receita_total,
        vendas_por_lista=[
            {"tipo": k, "vendas": v["vendas"], "receita": float(v["receita"])}
            for k, v in vendas_por_lista.items()
        ],
        vendas_por_promoter=[
            {"promoter": k, "vendas": v["vendas"], "receita": float(v["receita"])}
            for k, v in vendas_por_promoter.items()
        ]
    )

@router.get("/vendas/{evento_id}/csv")
async def exportar_vendas_csv(
    evento_id: int,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Exportar relatório de vendas em CSV"""
    
    evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not evento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento não encontrado"
        )
    
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acesso negado"
        )
    
    transacoes = db.query(Transacao).filter(
        Transacao.evento_id == evento_id,
        Transacao.status == "aprovada"
    ).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        'ID Transação', 'CPF Comprador', 'Nome Comprador', 'Email', 'Telefone',
        'Valor', 'Método Pagamento', 'Lista', 'Promoter', 'Data Compra'
    ])
    
    for transacao in transacoes:
        lista = db.query(Lista).filter(Lista.id == transacao.lista_id).first()
        promoter_nome = ""
        if lista and lista.promoter_id:
            promoter = db.query(Usuario).filter(Usuario.id == lista.promoter_id).first()
            if promoter:
                promoter_nome = promoter.nome
        
        writer.writerow([
            transacao.id,
            transacao.cpf_comprador,
            transacao.nome_comprador,
            transacao.email_comprador or "",
            transacao.telefone_comprador or "",
            float(transacao.valor),
            transacao.metodo_pagamento or "",
            lista.nome if lista else "",
            promoter_nome,
            transacao.criado_em.strftime("%d/%m/%Y %H:%M:%S")
        ])
    
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=vendas_evento_{evento_id}.csv"}
    )

@router.get("/checkins/{evento_id}/csv")
async def exportar_checkins_csv(
    evento_id: int,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Exportar relatório de check-ins em CSV"""
    
    evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not evento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento não encontrado"
        )
    
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acesso negado"
        )
    
    checkins = db.query(Checkin).filter(Checkin.evento_id == evento_id).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        'ID Check-in', 'CPF', 'Nome', 'Método Check-in', 
        'Data Check-in', 'Responsável Check-in'
    ])
    
    for checkin in checkins:
        responsavel = ""
        if checkin.usuario_id:
            usuario = db.query(Usuario).filter(Usuario.id == checkin.usuario_id).first()
            if usuario:
                responsavel = usuario.nome
        
        writer.writerow([
            checkin.id,
            checkin.cpf,
            checkin.nome,
            checkin.metodo_checkin,
            checkin.checkin_em.strftime("%d/%m/%Y %H:%M:%S"),
            responsavel
        ])
    
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=checkins_evento_{evento_id}.csv"}
    )

@router.get("/auditoria")
async def exportar_logs_auditoria(
    data_inicio: Optional[date] = None,
    data_fim: Optional[date] = None,
    cpf_usuario: Optional[str] = None,
    evento_id: Optional[int] = None,
    formato: str = "json",
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(verificar_permissao_admin)
):
    """Exportar logs de auditoria (apenas admins)"""
    
    from ..models import LogAuditoria
    
    query = db.query(LogAuditoria)
    
    if data_inicio:
        query = query.filter(LogAuditoria.criado_em >= data_inicio)
    
    if data_fim:
        query = query.filter(LogAuditoria.criado_em <= data_fim)
    
    if cpf_usuario:
        query = query.filter(LogAuditoria.cpf_usuario == cpf_usuario)
    
    if evento_id:
        query = query.filter(LogAuditoria.evento_id == evento_id)
    
    logs = query.order_by(LogAuditoria.criado_em.desc()).limit(1000).all()
    
    if formato == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow([
            'ID', 'CPF Usuário', 'Ação', 'Tabela', 'Registro ID',
            'IP Origem', 'Status', 'Data/Hora', 'Detalhes'
        ])
        
        for log in logs:
            writer.writerow([
                log.id,
                log.cpf_usuario,
                log.acao,
                log.tabela_afetada or "",
                log.registro_id or "",
                log.ip_origem or "",
                log.status,
                log.criado_em.strftime("%d/%m/%Y %H:%M:%S"),
                log.detalhes or ""
            ])
        
        output.seek(0)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=auditoria.csv"}
        )
    
    else:
        return {
            "total": len(logs),
            "logs": [
                {
                    "id": log.id,
                    "cpf_usuario": log.cpf_usuario,
                    "acao": log.acao,
                    "tabela_afetada": log.tabela_afetada,
                    "registro_id": log.registro_id,
                    "ip_origem": log.ip_origem,
                    "status": log.status,
                    "criado_em": log.criado_em.isoformat(),
                    "detalhes": log.detalhes
                }
                for log in logs
            ]
        }

@router.get("/vendas/{evento_id}/excel")
async def exportar_vendas_excel(
    evento_id: int,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Exportar relatório de vendas em Excel"""
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not evento:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Vendas"
    
    headers = ['ID', 'CPF', 'Nome', 'Email', 'Telefone', 'Valor', 'Método', 'Lista', 'Promoter', 'Data', 'Status']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    transacoes = db.query(Transacao).join(Lista).outerjoin(Usuario).filter(
        Transacao.evento_id == evento_id,
        Transacao.status == "aprovada"
    ).all()
    
    for row, transacao in enumerate(transacoes, 2):
        ws.cell(row=row, column=1, value=transacao.id)
        ws.cell(row=row, column=2, value=transacao.cpf_comprador)
        ws.cell(row=row, column=3, value=transacao.nome_comprador)
        ws.cell(row=row, column=4, value=transacao.email_comprador)
        ws.cell(row=row, column=5, value=transacao.telefone_comprador)
        ws.cell(row=row, column=6, value=float(transacao.valor))
        ws.cell(row=row, column=7, value=transacao.metodo_pagamento)
        ws.cell(row=row, column=8, value=transacao.lista.nome if transacao.lista else "")
        ws.cell(row=row, column=9, value=transacao.promoter.nome if transacao.promoter else "")
        ws.cell(row=row, column=10, value=transacao.criado_em.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row, column=11, value=transacao.status.upper())
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=vendas_evento_{evento_id}.xlsx"}
    )

@router.get("/dashboard/export/{formato}")
async def exportar_dashboard(
    formato: str,
    evento_id: Optional[int] = None,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Exportar dados do dashboard em diferentes formatos"""
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    if formato not in ["pdf", "csv", "excel"]:
        raise HTTPException(status_code=400, detail="Formato não suportado")
    
    transacoes_query = db.query(Transacao).filter(Transacao.status == "aprovada")
    checkins_query = db.query(Checkin)
    
    if usuario_atual.tipo.value != "admin":
        transacoes_query = transacoes_query.join(Evento).filter(Evento.empresa_id == usuario_atual.empresa_id)
        checkins_query = checkins_query.join(Evento).filter(Evento.empresa_id == usuario_atual.empresa_id)
    
    if evento_id:
        transacoes_query = transacoes_query.filter(Transacao.evento_id == evento_id)
        checkins_query = checkins_query.filter(Checkin.evento_id == evento_id)
    
    if formato == "excel":
        wb = Workbook()
        
        ws_vendas = wb.active
        ws_vendas.title = "Vendas"
        headers_vendas = ['CPF', 'Nome', 'Valor', 'Data', 'Método', 'Status']
        
        for col, header in enumerate(headers_vendas, 1):
            cell = ws_vendas.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        transacoes = transacoes_query.all()
        for row, transacao in enumerate(transacoes, 2):
            ws_vendas.cell(row=row, column=1, value=transacao.cpf_comprador)
            ws_vendas.cell(row=row, column=2, value=transacao.nome_comprador)
            ws_vendas.cell(row=row, column=3, value=float(transacao.valor))
            ws_vendas.cell(row=row, column=4, value=transacao.criado_em.strftime("%d/%m/%Y"))
            ws_vendas.cell(row=row, column=5, value=transacao.metodo_pagamento)
            ws_vendas.cell(row=row, column=6, value=transacao.status)
        
        ws_checkins = wb.create_sheet("Check-ins")
        headers_checkins = ['CPF', 'Nome', 'Data Check-in', 'Método']
        
        for col, header in enumerate(headers_checkins, 1):
            cell = ws_checkins.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        checkins = checkins_query.all()
        for row, checkin in enumerate(checkins, 2):
            ws_checkins.cell(row=row, column=1, value=checkin.cpf)
            ws_checkins.cell(row=row, column=2, value=checkin.nome)
            ws_checkins.cell(row=row, column=3, value=checkin.checkin_em.strftime("%d/%m/%Y %H:%M"))
            ws_checkins.cell(row=row, column=4, value=checkin.metodo_checkin)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    
    elif formato == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['=== RELATÓRIO DASHBOARD ==='])
        writer.writerow(['Data:', datetime.now().strftime('%d/%m/%Y %H:%M')])
        writer.writerow([])
        
        writer.writerow(['=== VENDAS ==='])
        writer.writerow(['CPF', 'Nome', 'Valor', 'Data', 'Método', 'Status'])
        
        transacoes = transacoes_query.all()
        for transacao in transacoes:
            writer.writerow([
                transacao.cpf_comprador,
                transacao.nome_comprador,
                str(transacao.valor),
                transacao.criado_em.strftime('%d/%m/%Y'),
                transacao.metodo_pagamento,
                transacao.status
            ])
        
        writer.writerow([])
        writer.writerow(['=== CHECK-INS ==='])
        writer.writerow(['CPF', 'Nome', 'Data Check-in', 'Método'])
        
        checkins = checkins_query.all()
        for checkin in checkins:
            writer.writerow([
                checkin.cpf,
                checkin.nome,
                checkin.checkin_em.strftime('%d/%m/%Y %H:%M'),
                checkin.metodo_checkin
            ])
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=dashboard_{datetime.now().strftime('%Y%m%d')}.csv"}
        )
    
    return {"message": "Formato PDF em desenvolvimento"}

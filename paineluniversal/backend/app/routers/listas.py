from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Response
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from ..database import get_db
from ..models import Lista, Evento, Usuario, TipoLista, Transacao, Checkin
from ..schemas import (
    Lista as ListaSchema, ListaCreate, ListaDetalhada, 
    DashboardListas, ConvidadoCreate, ConvidadoImport
)
from ..auth import obter_usuario_atual
import uuid
import re
import csv
import io
from decimal import Decimal

router = APIRouter()

@router.post("/", response_model=ListaSchema)
async def criar_lista(
    lista: ListaCreate,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Criar nova lista para evento"""
    
    evento = db.query(Evento).filter(Evento.id == lista.evento_id).first()
    if not evento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento não encontrado"
        )
    
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acesso negado"
        )
    
    db_lista = Lista(**lista.dict())
    db.add(db_lista)
    db.commit()
    db.refresh(db_lista)
    
    return db_lista

@router.get("/evento/{evento_id}", response_model=List[ListaSchema])
async def listar_listas_evento(
    evento_id: int,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Listar listas de um evento"""
    
    evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not evento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento não encontrado"
        )
    
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acesso negado"
        )
    
    listas = db.query(Lista).filter(Lista.evento_id == evento_id).all()
    return listas

@router.get("/promoter/{promoter_id}", response_model=List[ListaSchema])
async def listar_listas_promoter(
    promoter_id: int,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Listar listas de um promoter"""
    
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.id != promoter_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acesso negado"
        )
    
    listas = db.query(Lista).filter(Lista.promoter_id == promoter_id).all()
    return listas

@router.put("/{lista_id}", response_model=ListaSchema)
async def atualizar_lista(
    lista_id: int,
    lista_update: ListaCreate,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Atualizar dados da lista"""
    
    lista = db.query(Lista).filter(Lista.id == lista_id).first()
    if not lista:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lista não encontrada"
        )
    
    evento = db.query(Evento).filter(Evento.id == lista.evento_id).first()
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acesso negado"
        )
    
    for field, value in lista_update.dict(exclude={'evento_id'}).items():
        setattr(lista, field, value)
    
    db.commit()
    db.refresh(lista)
    
    return lista

@router.delete("/{lista_id}")
async def desativar_lista(
    lista_id: int,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Desativar lista"""
    
    lista = db.query(Lista).filter(Lista.id == lista_id).first()
    if not lista:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lista não encontrada"
        )
    
    evento = db.query(Evento).filter(Evento.id == lista.evento_id).first()
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acesso negado"
        )
    
    lista.ativa = False
    db.commit()
    
    return {"mensagem": "Lista desativada com sucesso"}

@router.get("/detalhada/{lista_id}", response_model=ListaDetalhada)
async def obter_lista_detalhada(
    lista_id: int,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Obter lista com métricas detalhadas"""
    
    lista = db.query(Lista).filter(Lista.id == lista_id).first()
    if not lista:
        raise HTTPException(status_code=404, detail="Lista não encontrada")
    
    evento = db.query(Evento).filter(Evento.id == lista.evento_id).first()
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    total_convidados = db.query(Transacao).filter(
        Transacao.lista_id == lista_id,
        Transacao.status == "aprovada"
    ).count()
    
    convidados_presentes = db.query(Checkin).join(Transacao).filter(
        Transacao.lista_id == lista_id,
        Transacao.status == "aprovada"
    ).count()
    
    receita_gerada = db.query(func.sum(Transacao.valor)).filter(
        Transacao.lista_id == lista_id,
        Transacao.status == "aprovada"
    ).scalar() or Decimal('0.00')
    
    taxa_presenca = (convidados_presentes / total_convidados * 100) if total_convidados > 0 else 0
    
    promoter_nome = None
    if lista.promoter_id:
        promoter = db.query(Usuario).filter(Usuario.id == lista.promoter_id).first()
        promoter_nome = promoter.nome if promoter else None
    
    return ListaDetalhada(
        **lista.__dict__,
        total_convidados=total_convidados,
        convidados_presentes=convidados_presentes,
        taxa_presenca=round(taxa_presenca, 2),
        receita_gerada=receita_gerada,
        promoter_nome=promoter_nome
    )

@router.post("/{lista_id}/convidados/import")
async def importar_convidados(
    lista_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Importar convidados via CSV/Excel"""
    
    import pandas as pd
    
    lista = db.query(Lista).filter(Lista.id == lista_id).first()
    if not lista:
        raise HTTPException(status_code=404, detail="Lista não encontrada")
    
    evento = db.query(Evento).filter(Evento.id == lista.evento_id).first()
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    try:
        content = await file.read()
        
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.StringIO(content.decode('utf-8')))
        elif file.filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(io.BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Formato não suportado. Use CSV ou Excel.")
        
        required_columns = ['cpf', 'nome']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(
                status_code=400, 
                detail=f"Colunas obrigatórias ausentes: {', '.join(missing_columns)}"
            )
        
        convidados_criados = 0
        erros = []
        
        for index, row in df.iterrows():
            try:
                cpf = re.sub(r'\D', '', str(row['cpf']))
                if len(cpf) != 11:
                    erros.append(f"Linha {index + 2}: CPF inválido")
                    continue
                
                cpf_formatado = f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"
                
                transacao_existente = db.query(Transacao).filter(
                    Transacao.cpf_comprador == cpf_formatado,
                    Transacao.evento_id == evento.id
                ).first()
                
                if transacao_existente:
                    erros.append(f"Linha {index + 2}: CPF {cpf_formatado} já cadastrado no evento")
                    continue
                
                transacao_data = {
                    'cpf_comprador': cpf_formatado,
                    'nome_comprador': str(row['nome']),
                    'email_comprador': str(row.get('email', '')),
                    'telefone_comprador': str(row.get('telefone', '')),
                    'valor': lista.preco,
                    'status': 'aprovada',
                    'lista_id': lista_id,
                    'evento_id': evento.id,
                    'usuario_id': usuario_atual.id,
                    'codigo_transacao': str(uuid.uuid4()),
                    'qr_code_ticket': f"TICKET-{str(uuid.uuid4())[:8].upper()}-{evento.id}"
                }
                
                db_transacao = Transacao(**transacao_data)
                db.add(db_transacao)
                convidados_criados += 1
                
            except Exception as e:
                erros.append(f"Linha {index + 2}: {str(e)}")
        
        lista.vendas_realizadas += convidados_criados
        db.commit()
        
        return {
            "convidados_criados": convidados_criados,
            "total_linhas": len(df),
            "erros": erros
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")

@router.get("/{lista_id}/convidados/export/{formato}")
async def exportar_convidados(
    lista_id: int,
    formato: str,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Exportar convidados da lista em CSV/Excel"""
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    lista = db.query(Lista).filter(Lista.id == lista_id).first()
    if not lista:
        raise HTTPException(status_code=404, detail="Lista não encontrada")
    
    evento = db.query(Evento).filter(Evento.id == lista.evento_id).first()
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    convidados = db.query(Transacao).outerjoin(Checkin).filter(
        Transacao.lista_id == lista_id,
        Transacao.status == "aprovada"
    ).all()
    
    if formato == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = f"Lista {lista.nome}"
        
        headers = ['CPF', 'Nome', 'Email', 'Telefone', 'QR Code', 'Status Presença', 'Data Check-in']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for row, convidado in enumerate(convidados, 2):
            checkin = db.query(Checkin).filter(Checkin.transacao_id == convidado.id).first()
            
            ws.cell(row=row, column=1, value=convidado.cpf_comprador)
            ws.cell(row=row, column=2, value=convidado.nome_comprador)
            ws.cell(row=row, column=3, value=convidado.email_comprador)
            ws.cell(row=row, column=4, value=convidado.telefone_comprador)
            ws.cell(row=row, column=5, value=convidado.qr_code_ticket)
            ws.cell(row=row, column=6, value="Presente" if checkin else "Ausente")
            ws.cell(row=row, column=7, value=checkin.checkin_em.strftime("%d/%m/%Y %H:%M") if checkin else "")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=lista_{lista.nome}_{lista_id}.xlsx"}
        )
    
    elif formato == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['CPF', 'Nome', 'Email', 'Telefone', 'QR Code', 'Status Presença', 'Data Check-in'])
        
        for convidado in convidados:
            checkin = db.query(Checkin).filter(Checkin.transacao_id == convidado.id).first()
            
            writer.writerow([
                convidado.cpf_comprador,
                convidado.nome_comprador,
                convidado.email_comprador,
                convidado.telefone_comprador,
                convidado.qr_code_ticket,
                "Presente" if checkin else "Ausente",
                checkin.checkin_em.strftime("%d/%m/%Y %H:%M") if checkin else ""
            ])
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=lista_{lista.nome}_{lista_id}.csv"}
        )

@router.get("/dashboard/{evento_id}")
async def obter_dashboard_listas(
    evento_id: int,
    db: Session = Depends(get_db),
    usuario_atual: Usuario = Depends(obter_usuario_atual)
):
    """Dashboard de listas para um evento"""
    
    evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not evento:
        raise HTTPException(status_code=404, detail="Evento não encontrado")
    
    if (usuario_atual.tipo.value != "admin" and 
        usuario_atual.empresa_id != evento.empresa_id):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    listas = db.query(Lista).filter(Lista.evento_id == evento_id).all()
    total_listas = len(listas)
    
    total_convidados = db.query(Transacao).filter(
        Transacao.evento_id == evento_id,
        Transacao.status == "aprovada"
    ).count()
    
    total_presentes = db.query(Checkin).filter(
        Checkin.evento_id == evento_id
    ).count()
    
    taxa_presenca_geral = (total_presentes / total_convidados * 100) if total_convidados > 0 else 0
    
    listas_mais_ativas = []
    for lista in listas[:5]:
        convidados = db.query(Transacao).filter(
            Transacao.lista_id == lista.id,
            Transacao.status == "aprovada"
        ).count()
        listas_mais_ativas.append({
            "nome": lista.nome,
            "tipo": lista.tipo.value,
            "convidados": convidados
        })
    
    return DashboardListas(
        total_listas=total_listas,
        total_convidados=total_convidados,
        total_presentes=total_presentes,
        taxa_presenca_geral=round(taxa_presenca_geral, 2),
        listas_mais_ativas=listas_mais_ativas,
        promoters_destaque=[],
        convidados_por_tipo=[],
        presencas_tempo_real=[]
    )
